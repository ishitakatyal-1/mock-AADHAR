# -*- coding: utf-8 -*-
"""
Created on Mon Sep 23 09:23:21 2019

@author: Ishita
"""
import tkinter as tk
from tkinter import *
from openpyxl import *
import os

home = os.getcwd()
#print(home)

wb = load_workbook(home + "\\" + "MockAADHAR.xlsx")
sheet = wb.active

def excel_heading():
    sheet.cell(row=1, column=1).value = "Date of enrollment"
    sheet.cell(row=1, column=2).value = "Enrollment ID"
    sheet.cell(row=1, column=3).value = "AADHAR Number"
    sheet.cell(row=1, column=4).value = "Reference ID"
    sheet.cell(row=1, column=5).value = "Reference Number"
    sheet.cell(row=1, column=6).value = "Name"
    sheet.cell(row=1, column=7).value = "Date of Birth"
    sheet.cell(row=1, column=8).value = "Gender"
    sheet.cell(row=1, column=9).value = "Father's Name"
    sheet.cell(row=1, column=10).value = "Mother's Name"
    sheet.cell(row=1, column=11).value = "Address"
    sheet.cell(row=1, column=12).value = "Date of issue"
    wb.save(home + "\\" + "MockAADHAR.xlsx")
    
def clear_record():
    date_of_enrollment.delete(0, "end")
    enrollment_id.delete(0, "end")
    aadhar_number.delete(0, "end")
    reference_id.delete(0, "end")
    reference_number.delete(0, "end")
    name.delete(0, "end")
    date_of_birth.delete(0, "end")
    gender.delete(0, "end")
    father_name.delete(0, "end")
    mother_name.delete(0, "end")
    address.delete(0, "end")
    date_of_issue.delete(0, "end")
    
def insert_record():
    if(date_of_enrollment.get() == "" or
       enrollment_id.get() == "" or
       aadhar_number.get() == "" or
       reference_id.get() == "" or
       reference_number.get() == "" or
       name.get() == "" or
       date_of_birth.get() == "" or
       gender.get() == "" or
       father_name.get() == "" or
       mother_name.get() == "" or
       address.get() == "" or
       date_of_issue.get() == ""):
        print("Empty input not allowed.")
        
    else:
        current_row = sheet.max_row
        sheet.cell(row=current_row+1, column=1).value = date_of_enrollment.get()
        sheet.cell(row=current_row+1, column=2).value = enrollment_id.get()
        sheet.cell(row=current_row+1, column=3).value = aadhar_number.get()
        sheet.cell(row=current_row+1, column=4).value = reference_id.get()
        sheet.cell(row=current_row+1, column=5).value = reference_number.get()
        sheet.cell(row=current_row+1, column=6).value = name.get()
        sheet.cell(row=current_row+1, column=7).value = date_of_birth.get()
        sheet.cell(row=current_row+1, column=8).value = gender.get()
        sheet.cell(row=current_row+1, column=9).value = father_name.get()
        sheet.cell(row=current_row+1, column=10).value = mother_name.get()
        sheet.cell(row=current_row+1, column=11).value = address.get()
        sheet.cell(row=current_row+1, column=12).value = date_of_issue.get()
        
        #print(date_of_enrollment.get())
        enrollment_date = date_of_enrollment.get()
        #print(enrollment_id.get())
        enrollment_id_aadhar = enrollment_id.get()
        directory = os.path.join(home + "\\" + enrollment_date)        
        folder_name = os.path.join(directory + "\\" + enrollment_id_aadhar)
        if not os.path.exists(folder_name):
            if not os.path.exists(directory):
                os.makedirs(directory)
                print("Directory created.")
                os.makedirs(folder_name)
                print("Folder created")
                wb.save(home + "\\" + "MockAADHAR.xlsx")
            else:
                os.makedirs(folder_name)
                print("Folder created")
                wb.save(home + "\\" + "MockAADHAR.xlsx")
        else:
            print("Record already exists.")
        
        clear_record()
                
mock_aadhar = tk.Tk()
mock_aadhar.title("Mock AADHAR")
mock_aadhar.config(background="white")
excel_heading()
        
mainframe1 = tk.Frame(mock_aadhar)
mainframe1.pack()

label00 = tk.Label(mainframe1, text="AADHAR Registration Form", font="Times 16")
label01 = tk.Label(mainframe1, text="Date of enrollment: ", font="Times 16")
label02 = tk.Label(mainframe1, text="Enrollment ID: ", font="Times 16")
label03 = tk.Label(mainframe1, text="AADHAR Number: ", font="Times 16")
label04 = tk.Label(mainframe1, text="Reference ID: ", font="Times 16")
label05 = tk.Label(mainframe1, text="Reference Number: ", font="Times 16")
label06 = tk.Label(mainframe1, text="Name: ", font="Times 16")
label07 = tk.Label(mainframe1, text="Date of birth: ", font="Times 16")
label08 = tk.Label(mainframe1, text="Gender: ", font="Times 16")
label09 = tk.Label(mainframe1, text="Father's Name: ", font="Times 16")
label10 = tk.Label(mainframe1, text="Mother's Name: ", font="Times 16")
label11 = tk.Label(mainframe1, text="Address: ", font="Times 16")
label12 = tk.Label(mainframe1, text="Date of Issue: ", font="Times 16")

date_of_enrollment = tk.Entry(mainframe1)
enrollment_id = tk.Entry(mainframe1)
aadhar_number = tk.Entry(mainframe1)
reference_id = tk.Entry(mainframe1)
reference_number = tk.Entry(mainframe1)
name = tk.Entry(mainframe1)
date_of_birth = tk.Entry(mainframe1)
gender = tk.Entry(mainframe1)
father_name = tk.Entry(mainframe1)
mother_name = tk.Entry(mainframe1)
address = tk.Entry(mainframe1)
date_of_issue = tk.Entry(mainframe1)

button_save_record = tk.Button(mainframe1, text="Save Record", font="Times 14", command=insert_record)
button_quit = tk.Button(mainframe1, text="Quit", font="Times 14", command=mock_aadhar.destroy)

label00.grid(row=0, column=0, sticky=tk.W, padx=5)
label01.grid(row=1, column=0, sticky=tk.W, padx=5)
label02.grid(row=2, column=0, sticky=tk.W, padx=5)
label03.grid(row=3, column=0, sticky=tk.W, padx=5)
label04.grid(row=4, column=0, sticky=tk.W, padx=5)
label05.grid(row=5, column=0, sticky=tk.W, padx=5)
label06.grid(row=6, column=0, sticky=tk.W, padx=5)
label07.grid(row=7, column=0, sticky=tk.W, padx=5)
label08.grid(row=8, column=0, sticky=tk.W, padx=5)
label09.grid(row=9, column=0, sticky=tk.W, padx=5)
label10.grid(row=10, column=0, sticky=tk.W, padx=5)
label11.grid(row=11, column=0, sticky=tk.W, padx=5)
label12.grid(row=12, column=0, sticky=tk.W, padx=5)

date_of_enrollment.grid(row=1, column=1, columnspan=2, sticky=tk.W)
enrollment_id.grid(row=2, column=1, columnspan=2, sticky=tk.W)
aadhar_number.grid(row=3, column=1, columnspan=2, sticky=tk.W)
reference_id.grid(row=4, column=1, columnspan=2, sticky=tk.W)
reference_number.grid(row=5, column=1, columnspan=2, sticky=tk.W)
name.grid(row=6, column=1, columnspan=2, sticky=tk.W)
date_of_birth.grid(row=7, column=1, columnspan=2, sticky=tk.W)
gender.grid(row=8, column=1, columnspan=2, sticky=tk.W)
father_name.grid(row=9, column=1, columnspan=2, sticky=tk.W)
mother_name.grid(row=10, column=1, columnspan=2, sticky=tk.W)
address.grid(row=11, column=1, columnspan=2, sticky=tk.W)
date_of_issue.grid(row=12, column=1, columnspan=2, sticky=tk.W)

button_save_record.grid(row=13, column=0,  sticky=tk.W)
button_quit.grid(row=13, column=1, sticky=tk.W)

mock_aadhar.mainloop()        
        
        
